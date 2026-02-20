import openpyxl
from openpyxl.utils import get_column_letter
import os
import glob
from datetime import datetime, timedelta
import pandas as pd
import traceback
import difflib

# --- CONFIGURATION ---
MOVEMENT_FOLDER = r'C:\Users\ADMIN\Desktop\Christopher\Arrears Reports'
BRANCH_DATA_FOLDER = r'C:\Users\ADMIN\Desktop\Christopher\Arrears Reports\Arears Reports formating folder\Documents'

def get_date_suffix(day):
    if 11 <= day <= 13: return 'th'
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')

def format_custom_date(dt, long_month=False, include_year=False):
    if not isinstance(dt, datetime): return str(dt)
    suffix = get_date_suffix(dt.day)
    month_fmt = "%B" if long_month else "%b"
    year_fmt = f" {dt.year}" if include_year else ""
    return f"{dt.day}{suffix} {dt.strftime(month_fmt)}{year_fmt}"

def get_latest_movement_report():
    files = glob.glob(os.path.join(MOVEMENT_FOLDER, "Arrears Movement Report *.xlsm"))
    if not files: return None
    files.sort() 
    return files[-1]

def get_all_branch_data():
    all_totals = {}
    files = [f for f in os.listdir(BRANCH_DATA_FOLDER) if f.lower().endswith(('.xlsx', '.csv'))]
    files = [f for f in files if "Movement Report" not in f]
    
    for filename in files:
        branch_id = filename.split()[0].lower() if ' ' in filename else filename.split('.')[0].lower()
        f_path = os.path.join(BRANCH_DATA_FOLDER, filename)
        
        try:
            df = pd.read_csv(f_path, header=None, encoding='latin1').fillna('') if filename.endswith('.csv') else pd.read_excel(f_path, header=None).fillna('')
            officer_starts = [i for i in range(len(df)) if "Loan Officer" in str(df.iloc[i, 0])]
            for start_row in officer_starts:
                name = " ".join(str(df.iloc[start_row, 1]).split()).lower().strip()
                for j in range(start_row + 1, len(df)):
                    if str(df.iloc[j, 0]).strip() == '' and str(df.iloc[j, 1]).strip() != '':
                        try:
                            val = float(df.iloc[j, 15]) 
                            all_totals[name] = {"val": val, "branch": branch_id}
                            break
                        except: continue
        except Exception as e:
            print(f"   ❌ Error reading {filename}: {e}")
    return all_totals

def run_daily_automation():
    print("--- 🏁 STARTING DYNAMIC UPDATE ---")
    today_obj = datetime.now()
    yesterday_obj = today_obj - timedelta(days=2 if today_obj.weekday() == 0 else 1)
    
    latest_report = get_latest_movement_report()
    if not latest_report: return

    new_report_path = os.path.join(MOVEMENT_FOLDER, f"Arrears Movement Report {today_obj.strftime('%Y.%m.%d')}.xlsm")
    wb = openpyxl.load_workbook(latest_report, keep_vba=True)
    ws = wb.active 

    # 1. FIND HEADERS
    var_col_idx = None
    header_row_idx = None
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            if "Var." in str(ws.cell(row=r, column=c).value):
                var_col_idx, header_row_idx = c, r
                break
    
    today_col, yesterday_col = var_col_idx - 1, var_col_idx - 2
    jan_col_letter = "L"

    # 2. UPDATE HEADERS (FIXED SECTION)
    ws.cell(row=header_row_idx, column=yesterday_col).value = format_custom_date(yesterday_obj, include_year=True)
    ws.cell(row=header_row_idx, column=today_col).value = format_custom_date(today_obj, include_year=True)
    
    # Daily Variance Header
    ws.cell(row=header_row_idx, column=var_col_idx).value = f"Var. {format_custom_date(today_obj)} & {format_custom_date(yesterday_obj)} {today_obj.year}"
    
    # Static Reference Variance Header (Fixed 2nd Jan 2026)
    static_ref_date = "2nd January 2026"
    ws.cell(row=header_row_idx, column=var_col_idx + 1).value = f"Var. {format_custom_date(today_obj)} & {static_ref_date}"

    # 3. DATA PROCESSING
    branch_data = get_all_branch_data()
    found_in_master = set()

    for row in range(header_row_idx + 1, ws.max_row + 1):
        target_cell = ws.cell(row=row, column=today_col)
        lo_name_raw = ws.cell(row=row, column=2).value
        
        if isinstance(target_cell.value, str) and str(target_cell.value).startswith('='):
            continue 

        ws.cell(row=row, column=yesterday_col).value = target_cell.value
        
        if lo_name_raw:
            m_name = " ".join(str(lo_name_raw).split()).lower().strip()
            matches = difflib.get_close_matches(m_name, branch_data.keys(), n=1, cutoff=0.8)
            if matches:
                match_name = matches[0]
                target_cell.value = branch_data[match_name]["val"]
                found_in_master.add(match_name)
            else:
                target_cell.value = 0

        # Update Formulas
        t_let, y_let = get_column_letter(today_col), get_column_letter(yesterday_col)
        ws.cell(row=row, column=var_col_idx).value = f"={t_let}{row}-{y_let}{row}"
        ws.cell(row=row, column=var_col_idx + 1).value = f"={t_let}{row}-{jan_col_letter}{row}"

    # Second Pass: Add New Officers
    new_officers = [name for name in branch_data.keys() if name not in found_in_master]
    for new_name in new_officers:
        branch_tag = branch_data[new_name]["branch"]
        print(f"➕ New Officer Detected: {new_name} ({branch_tag})")
        
        for r in range(header_row_idx + 1, ws.max_row + 1):
            cell_val = str(ws.cell(row=r, column=2).value).lower()
            if branch_tag in cell_val and "total" in cell_val:
                ws.insert_rows(r)
                ws.cell(row=r, column=2).value = new_name.title()
                ws.cell(row=r, column=yesterday_col).value = 0
                ws.cell(row=r, column=today_col).value = branch_data[new_name]["val"]
                
                t_let, y_let = get_column_letter(today_col), get_column_letter(yesterday_col)
                ws.cell(row=r, column=var_col_idx).value = f"={t_let}{r}-{y_let}{r}"
                ws.cell(row=r, column=var_col_idx + 1).value = f"={t_let}{r}-{jan_col_letter}{r}"
                break

    wb.save(new_report_path)
    print(f"\n✨ DONE! Saved: {os.path.basename(new_report_path)}")

if __name__ == "__main__":
    run_daily_automation()
    input("\nPress ENTER to close...")